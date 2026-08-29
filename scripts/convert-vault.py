# -*- coding: utf-8 -*-
"""密码薄 xlsx -> CSV 转换（供 sync-vault.ps1 调用）"""
import openpyxl
import csv
import io
import hashlib
import sys

SRC = r'D:\个人信息\密码薄.xlsx'
OUT = r'D:\Works\ZipperAgentMemory\memory\structured\password-vault.csv'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = [[str(c) if c is not None else '' for c in row] for row in ws.iter_rows(values_only=True)]

buf = io.StringIO()
w = csv.writer(buf)
w.writerows(rows)
content = buf.getvalue()

with io.open(OUT, 'w', newline='', encoding='utf-8-sig') as f:
    f.write(content)

print(f"rows={len(rows)} sha256={hashlib.sha256(content.encode('utf-8')).hexdigest()[:16]}")
