# -*- coding: utf-8 -*-
"""重新解析密码薄.xlsx（忽略图片），先读旧版对比，再写新 CSV"""
import openpyxl
import csv
import io
import os

SRC = r'D:\个人信息\密码薄.xlsx'
CSV_PATH = r'D:\Works\ZipperAgentMemory\memory\structured\password-vault.csv'

# 1. 先读旧 CSV（如果存在）
old_rows = []
if os.path.exists(CSV_PATH):
    with io.open(CSV_PATH, 'r', encoding='utf-8-sig') as f:
        old_rows = list(csv.reader(f))
    print(f'old rows (incl header): {len(old_rows)}')

# 2. 解析 xlsx
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = [[str(c) if c is not None else '' for c in row] for row in ws.iter_rows(values_only=True)]
print(f'new rows (incl header): {len(rows)}')
print('header:', rows[0])
print('images in sheet:', len(ws._images))

# 3. 对比
if old_rows:
    old_data = {tuple(r[:3]): r for r in old_rows[1:]}
    new_data = {tuple(r[:3]): r for r in rows[1:]}
    added = [r for k, r in new_data.items() if k not in old_data]
    removed = [r for k, r in old_data.items() if k not in new_data]
    changed = [r for k, r in new_data.items() if k in old_data and r != old_data[k]]
    print(f'\n== 新增 {len(added)} 条 ==')
    for r in added:
        print('  +', r[:4])
    print(f'\n== 删除 {len(removed)} 条 ==')
    for r in removed:
        print('  -', r[:4])
    print(f'\n== 修改 {len(changed)} 条 ==')
    for r in changed:
        print('  ~', r[:4])

# 4. 写新 CSV
with io.open(CSV_PATH, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerows(rows)
print(f'\nwritten: {CSV_PATH} ({len(rows)} rows)')
